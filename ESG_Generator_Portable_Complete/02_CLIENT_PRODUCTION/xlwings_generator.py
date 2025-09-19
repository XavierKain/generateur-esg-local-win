#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module de génération de questionnaires ESG avec xlwings
Supporte macOS (sandbox Excel) et Windows (répertoire temporaire) pour préserver 100% du formatage conditionnel
"""

import xlwings as xw
import os
import shutil
import pandas as pd
from pathlib import Path
import tempfile
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook

class XLWingsGenerator:
    """
    Générateur de questionnaires ESG utilisant xlwings
    Support Windows et macOS avec détection automatique du système
    """
    
    def __init__(self):
        """Initialiser le générateur xlwings"""
        # Détection automatique du système
        import platform
        
        if platform.system() == "Windows":
            # Windows - utiliser le répertoire temporaire système
            self.temp_dir = Path(tempfile.gettempdir()) / "xlwings_esg_temp"
            self.is_windows = True
        else:
            # macOS - répertoire sandbox Excel
            self.excel_sandbox_dir = Path.home() / "Library/Containers/com.microsoft.Excel/Data"
            self.temp_dir = self.excel_sandbox_dir / "temp_questionnaires"
            self.is_windows = False
        
        # Créer le répertoire temporaire s'il n'existe pas
        try:
            self.temp_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            print(f"Avertissement - Création répertoire temp: {e}")
        
        self.is_available_cached = None
        self.status_message = None
    
    def is_available(self) -> Tuple[bool, str]:
        """
        Vérifie si xlwings et Excel sont disponibles
        """
        if self.is_available_cached is not None:
            return self.is_available_cached, self.status_message
        
        try:
            # 1. Vérifier si xlwings est installé
            import xlwings as xw
            
            # 2. Vérifications spécifiques selon le système
            if self.is_windows:
                # Windows - vérification simplifiée
                if not self.temp_dir.parent.exists():
                    self.is_available_cached = False
                    self.status_message = f"Répertoire temporaire non accessible: {self.temp_dir.parent}"
                    return False, self.status_message
            else:
                # macOS - vérification sandbox
                if not self.excel_sandbox_dir.exists():
                    self.is_available_cached = False
                    self.status_message = f"Répertoire Excel sandbox non trouvé: {self.excel_sandbox_dir}"
                    return False, self.status_message
            
            # 3. Tester si on peut écrire dans le répertoire
            test_file = self.temp_dir / "test_write.txt"
            try:
                test_file.write_text("test")
                test_file.unlink()
            except Exception as e:
                self.is_available_cached = False
                self.status_message = f"Pas d'accès en écriture: {e}"
                return False, self.status_message
            
            # 4. Tester xlwings avec Excel (avec logs détaillés)
            try:
                print(f"🔍 [XLWings Debug] Test système: {'Windows' if self.is_windows else 'macOS'}")
                
                if self.is_windows:
                    # Sur Windows - diagnostic approfondi
                    print("🔍 [XLWings Debug] Vérification xw.apps...")
                    print(f"🔍 [XLWings Debug] xw.apps = {xw.apps}")
                    print(f"🔍 [XLWings Debug] type(xw.apps) = {type(xw.apps)}")
                    
                    if xw.apps is None:
                        self.is_available_cached = False
                        self.status_message = "XLWings apps collection est None - Excel non initialisé"
                        print(f"❌ [XLWings Debug] {self.status_message}")
                        return False, self.status_message
                    
                    print("🔍 [XLWings Debug] Tentative création App...")
                    
                    # Tentative 1: Méthode standard
                    try:
                        app = xw.App(visible=False, add_book=False)
                        print(f"✅ [XLWings Debug] App créée: {app}")
                        
                        # Test des propriétés
                        print(f"🔍 [XLWings Debug] App.books: {app.books}")
                        print(f"🔍 [XLWings Debug] App.version: {getattr(app, 'version', 'N/A')}")
                        
                        app.quit()
                        print("✅ [XLWings Debug] App fermée avec succès")
                        
                    except Exception as app_error:
                        print(f"❌ [XLWings Debug] Erreur App standard: {app_error}")
                        print(f"🔍 [XLWings Debug] Type erreur: {type(app_error)}")
                        
                        # Tentative 2: Avec COM direct
                        try:
                            print("🔍 [XLWings Debug] Tentative COM direct...")
                            import win32com.client
                            xl_com = win32com.client.Dispatch('Excel.Application')
                            xl_com.Visible = False
                            print(f"✅ [XLWings Debug] COM Excel OK: {xl_com.Version}")
                            xl_com.Quit()
                            
                            # Retry XLWings après COM
                            app = xw.App(visible=False)
                            app.quit()
                            print("✅ [XLWings Debug] XLWings OK après COM")
                            
                        except Exception as com_error:
                            print(f"❌ [XLWings Debug] COM échoué: {com_error}")
                            raise app_error  # Relancer l'erreur originale
                    
                else:
                    # macOS - méthode originale avec logs
                    print("🔍 [XLWings Debug] macOS - test standard")
                    app = xw.App(visible=False)
                    app.quit()
                    print("✅ [XLWings Debug] macOS OK")
                    
            except Exception as e:
                self.is_available_cached = False
                error_detail = f"Excel non accessible via xlwings: {e}"
                print(f"❌ [XLWings Debug] Erreur finale: {error_detail}")
                print(f"🔍 [XLWings Debug] Stack trace complet:")
                import traceback
                traceback.print_exc()
                
                self.status_message = error_detail
                return False, self.status_message
            
            self.is_available_cached = True
            if self.is_windows:
                self.status_message = "xlwings et Excel disponibles sur Windows"
            else:
                self.status_message = "xlwings et Excel disponibles dans le sandbox"
            return True, self.status_message
            
        except ImportError:
            self.is_available_cached = False
            self.status_message = "xlwings non installé"
            return False, self.status_message
        except Exception as e:
            self.is_available_cached = False
            self.status_message = f"Erreur xlwings: {str(e)}"
            return False, self.status_message
    
    def generate_single_questionnaire(self, template_path: str, data_row: Any, original_output_path: str) -> Dict[str, Any]:
        """
        Génère un questionnaire en utilisant xlwings dans le sandbox Excel
        
        Args:
            template_path: Chemin vers le template Excel
            data_row: Données pour remplir le questionnaire (Series pandas ou dict)
            original_output_path: Chemin de sortie souhaité
            
        Returns:
            Dict avec success, message, output_path, etc.
        """
        app = None
        sandbox_template = None
        sandbox_output = None
        
        try:
            # 1. Vérifier la disponibilité
            available, status = self.is_available()
            if not available:
                return {
                    "success": False,
                    "error": f"xlwings non disponible: {status}",
                    "output_path": None,
                    "method": "xlwings"
                }
            
            # 2. Copier le template dans le répertoire autorisé
            template_name = f"template_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
            sandbox_template = self.temp_dir / template_name
            shutil.copy2(template_path, sandbox_template)
            
            # 3. Créer le nom de fichier de sortie dans le sandbox
            output_name = f"output_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
            sandbox_output = self.temp_dir / output_name
            
            # 4. Ouvrir avec xlwings (Excel)
            app = xw.App(visible=False)
            wb = app.books.open(str(sandbox_template))
            
            # 5. Mettre à jour les données selon votre structure
            self._update_questionnaire_data(wb, data_row)
            
            # 6. Sauvegarder dans le sandbox
            wb.save(str(sandbox_output))
            wb.close()
            
            # 7. Copier le résultat vers la destination finale
            os.makedirs(os.path.dirname(original_output_path), exist_ok=True)
            shutil.copy2(sandbox_output, original_output_path)
            
            return {
                "success": True,
                "message": "Questionnaire généré avec xlwings (formatage 100% préservé)",
                "output_path": original_output_path,
                "method": "xlwings",
                "sandbox_used": True
            }
            
        except Exception as e:
            error_msg = f"Erreur xlwings: {str(e)}"
            print(error_msg)
            return {
                "success": False,
                "error": error_msg,
                "output_path": None,
                "method": "xlwings"
            }
        finally:
            # 8. Nettoyer les ressources
            if app:
                try:
                    app.quit()
                except:
                    pass
            
            # 9. Nettoyer les fichiers temporaires
            try:
                if sandbox_template and sandbox_template.exists():
                    sandbox_template.unlink()
                if sandbox_output and sandbox_output.exists():
                    sandbox_output.unlink()
            except Exception as e:
                print(f"Avertissement - Nettoyage fichiers temp: {e}")
    
    def _update_questionnaire_data(self, wb, data_row):
        """
        Met à jour les données dans le questionnaire Excel
        Utilise la même logique que la fonction openpyxl classique
        
        Args:
            wb: Workbook xlwings
            data_row: Données à insérer (dict avec les mappings colonnes/valeurs)
        """
        try:
            # Importer les formules template depuis le module principal
            from generateur_2025_streamlit import TEMPLATE_FORMULAS, extract_cell_references
            
            # 1. D'abord, remplir les données spéciales (locataire, adresse, etc.)
            if 'Questionnaire ESG' in [s.name for s in wb.sheets]:
                ws_questionnaire = wb.sheets['Questionnaire ESG']
                try:
                    # Locataire en B11
                    if 'nom_locataire' in data_row and data_row['nom_locataire']:
                        ws_questionnaire.range('B11').value = data_row['nom_locataire']
                    
                    # Adresse combinée avec ville en B8 - Logique Python
                    # Combine les champs 'ville' et 'adresse' avec ' - '
                    ville = data_row.get('ville', '').strip() if data_row.get('ville') else ''
                    adresse = data_row.get('adresse', '').strip() if data_row.get('adresse') else ''
                    
                    if ville and adresse:
                        adresse_combinee = f"{ville} - {adresse}"
                    elif ville:
                        adresse_combinee = ville
                    elif adresse:
                        adresse_combinee = adresse
                    else:
                        adresse_combinee = ''
                    
                    if adresse_combinee:
                        ws_questionnaire.range('B8').value = adresse_combinee
                        print(f"📍 [Debug] B8 mis à jour avec: '{adresse_combinee}'")
                    else:
                        print("⚠️ [Debug] Aucune donnée ville/adresse disponible pour B8")
                
                    
                    # Réponse certifiée en A22
                    if 'reponse_certifiee' in data_row and data_row['reponse_certifiee']:
                        ws_questionnaire.range('A22').value = data_row['reponse_certifiee']
                        
                except Exception as e:
                    print(f"Avertissement - Données spéciales: {e}")
            
            # 2. Utiliser le système de mapping des formules (TEMPLATE_FORMULAS)
            updates_count = 0
            for col_letter, formula in TEMPLATE_FORMULAS.items():
                cell_info = extract_cell_references(formula)
                
                if cell_info:
                    sheet_name, cell_ref = cell_info
                    try:
                        # Lire la valeur depuis les données préparées
                        data_key = f'col_{col_letter}'
                        if data_key in data_row:
                            source_value = data_row[data_key]
                            
                            if source_value is not None and str(source_value).strip() != '':
                                # Écrire dans le questionnaire
                                if sheet_name in [s.name for s in wb.sheets]:
                                    ws_target = wb.sheets[sheet_name]
                                    ws_target.range(cell_ref).value = source_value
                                    updates_count += 1
                            
                    except Exception as e:
                        print(f"Avertissement - Erreur cellule {col_letter}: {e}")
                        continue
            
            print(f"✅ xlwings: {updates_count} valeurs mises à jour")
                        
        except Exception as e:
            print(f"Erreur lors de la mise à jour des données: {e}")
            raise
    
    def _update_sheet_data(self, sheet, data_dict: Dict, sheet_name: str):
        """
        Met à jour les données d'une feuille spécifique
        Utilise la même logique que la fonction openpyxl
        
        Args:
            sheet: Feuille xlwings
            data_dict: Données 
            sheet_name: Nom de la feuille
        """
        try:
            # Pour l'instant, on utilise la logique des TEMPLATE_FORMULAS
            # qui couvre déjà les principales feuilles
            pass
                
        except Exception as e:
            print(f"Erreur mise à jour feuille {sheet_name}: {e}")

    def generate_multiple_questionnaires(self, template_path: str, data_list: List[Any], output_directory: str) -> List[Dict[str, Any]]:
        """
        Génère plusieurs questionnaires
        
        Args:
            template_path: Chemin vers le template
            data_list: Liste des données (Series pandas ou dicts)
            output_directory: Répertoire de sortie
            
        Returns:
            Liste des résultats de génération
        """
        results = []
        
        for i, data_row in enumerate(data_list):
            # Créer le nom de fichier de sortie
            filename = self._create_filename(data_row, i)
            output_path = Path(output_directory) / filename
            
            # Générer le questionnaire
            result = self.generate_single_questionnaire(template_path, data_row, str(output_path))
            result['index'] = i
            result['filename'] = filename
            results.append(result)
        
        return results
    
    def generate_questionnaires_to_zip(self, bdd_file, year, template_file, selected_indices, progress_callback=None):
        """
        Génère les questionnaires sélectionnés vers un ZIP avec xlwings
        Compatible avec l'interface Streamlit existante
        
        Args:
            bdd_file: Fichier BDD Excel
            year: Année à traiter 
            template_file: Fichier template
            selected_indices: Liste des indices des questionnaires à générer
            progress_callback: Callback pour la progression
            
        Returns:
            dict: Résultat avec zip_data et statistiques
        """
        import zipfile
        import io
        import tempfile
        from openpyxl import load_workbook
        
        try:
            # Lire les données de l'année depuis le fichier BDD (avec openpyxl pour la lecture)
            all_data = self._read_year_data_from_bdd(bdd_file, year)
            
            if not all_data:
                return {'success': False, 'error': f'Aucune donnée trouvée pour l\'année {year}'}
            
            # Filtrer selon les indices sélectionnés
            selected_data = [all_data[i] for i in selected_indices if i < len(all_data)]
            
            if not selected_data:
                return {'success': False, 'error': 'Aucun questionnaire sélectionné valide'}
            
            # Créer un buffer ZIP en mémoire
            zip_buffer = io.BytesIO()
            generated_count = 0
            failed_count = 0
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for i, data_row in enumerate(selected_data):
                    if progress_callback:
                        locataire = data_row.get('locataire', f'Questionnaire {i+1}')
                        progress_callback(i, len(selected_data), f"Génération {locataire}")
                    
                    try:
                        # Créer un fichier temporaire pour ce questionnaire
                        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                            temp_path = temp_file.name
                        
                        # Utiliser xlwings pour générer le questionnaire
                        result = self.generate_single_questionnaire(template_file, data_row, temp_path)
                        
                        if result.get('success', False):
                            # Créer la structure de dossiers
                            ville = data_row.get('ville', '')
                            adresse = data_row.get('adresse', '')
                            locataire = data_row.get('locataire', '')
                            folder_structure = f"{ville} - {adresse} - {locataire}"
                            filename = self._create_zip_filename(data_row, year, folder_structure)
                            
                            # Ajouter au ZIP
                            zipf.write(temp_path, filename)
                            generated_count += 1
                        else:
                            failed_count += 1
                        
                        # Nettoyer le fichier temporaire
                        try:
                            os.unlink(temp_path)
                        except:
                            pass
                            
                    except Exception as e:
                        failed_count += 1
                        print(f"Erreur xlwings pour {data_row.get('locataire', 'Unknown')}: {e}")
            
            # Progression finale
            if progress_callback:
                progress_callback(len(selected_data), len(selected_data), "Génération terminée avec xlwings")
            
            # Retourner les données ZIP
            zip_data = zip_buffer.getvalue()
            
            return {
                'success': True,
                'zip_data': zip_data,
                'generated_count': generated_count,
                'failed_count': failed_count,
                'total_processed': len(selected_data),
                'total_available': len(all_data)
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def _read_year_data_from_bdd(self, bdd_file, year):
        """Lit les données d'une année depuis le fichier BDD"""
        try:
            wb = load_workbook(bdd_file, data_only=True)
            if year not in wb.sheetnames:
                return []
                
            ws = wb[year]
            data = []
            
            for row in range(10, ws.max_row + 1):  # Commence ligne 10
                # Colonne O = nom du dossier (colonne 15)
                folder_name = ws.cell(row=row, column=15).value
                if not folder_name or str(folder_name).strip() == '':
                    continue
                
                # Extraire les données selon votre structure
                data_row = {
                    'numero_identification': ws.cell(row=row, column=1).value,
                    'locataire': ws.cell(row=row, column=6).value,
                    'adresse': ws.cell(row=row, column=4).value,
                    'ville': ws.cell(row=row, column=5).value,
                    'folder_name': folder_name,
                    'row_number': row
                }
                
                # Ajouter toutes les colonnes pour compatibilité
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    data_row[f'col_{col}'] = cell_value
                
                data.append(data_row)
            
            wb.close()
            return data
            
        except Exception as e:
            print(f"Erreur lecture BDD: {e}")
            return []
    
    def _create_zip_filename(self, data_row, year, folder_structure):
        """Crée le nom de fichier pour le ZIP"""
        locataire = data_row.get('locataire', 'Unknown')
        safe_locataire = "".join(c for c in locataire if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f"{folder_structure}/ESG_{safe_locataire}_{year}.xlsx"
        return filename
    
    def _create_filename(self, data_row: Any, index: int) -> str:
        """
        Crée le nom de fichier selon votre format
        
        Args:
            data_row: Données du questionnaire
            index: Index dans la liste
            
        Returns:
            Nom de fichier formaté
        """
        try:
            # Convertir en dict si c'est une Series pandas
            if hasattr(data_row, 'to_dict'):
                data_dict = data_row.to_dict()
            else:
                data_dict = dict(data_row) if data_row else {}
            
            # Extraire les informations pour le nom de fichier
            date_str = data_dict.get('date', datetime.now().strftime('%Y %m %d'))
            locataire = data_dict.get('nom_locataire', f'Locataire_{index}')
            adresse = data_dict.get('adresse', f'Adresse_{index}')
            
            # Format: "AAAA MM JJ – Questionnaire ESG – Rue, n° - locataire - #id"
            filename = f"{date_str} – Questionnaire ESG – {adresse} - {locataire} - #{index+1}.xlsx"
            
            # Nettoyer le nom de fichier pour éviter les caractères problématiques
            filename = "".join(c for c in filename if c.isalnum() or c in ' -_#.àéèêôûçÀÉÈÊÔÛÇ').strip()
            
            # Limiter la longueur
            if len(filename) > 100:
                filename = filename[:97] + "...xlsx"
                
            return filename
            
        except Exception as e:
            print(f"Erreur création nom fichier: {e}")
            return f"Questionnaire_ESG_{index+1}.xlsx"
    
    def cleanup_temp_files(self):
        """Nettoie les fichiers temporaires du sandbox"""
        try:
            if self.temp_dir.exists():
                for file in self.temp_dir.glob("*"):
                    try:
                        file.unlink()
                    except Exception as e:
                        print(f"Erreur suppression {file}: {e}")
                print(f"Nettoyage terminé: {self.temp_dir}")
        except Exception as e:
            print(f"Erreur nettoyage global: {e}")
    
    def get_info(self) -> Dict[str, Any]:
        """
        Retourne les informations sur le générateur xlwings
        
        Returns:
            Dict avec les informations de configuration
        """
        available, status = self.is_available()
        
        return {
            "available": available,
            "status": status,
            "excel_sandbox_dir": str(self.excel_sandbox_dir),
            "temp_dir": str(self.temp_dir),
            "temp_dir_exists": self.temp_dir.exists(),
            "method": "xlwings",
            "formatage_preservation": "100%",
            "requires_excel": True,
            "platform": "macOS"
        }

# Test de base du module
if __name__ == "__main__":
    # Test d'initialisation
    generator = XLWingsGenerator()
    available, status = generator.is_available()
    
    print(f"xlwings disponible: {available}")
    print(f"Statut: {status}")
    print(f"Informations: {generator.get_info()}")