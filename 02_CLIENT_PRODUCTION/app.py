#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
G√âN√âRATEUR ESG SIMPLIFI√â - Interface Streamlit
Application simplifi√©e pour la g√©n√©ration de questionnaires ESG uniquement
"""

import streamlit as st
import os
import tempfile
import shutil
import datetime
import time
import glob
import zipfile
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import warnings
from xlwings_generator import XLWingsGenerator

# Configuration portable
try:
    from config_portable import setup_portable_environment, get_recommended_generator, PORTABLE_MODE
    setup_portable_environment()
    FORCE_OPENPYXL = PORTABLE_MODE
except ImportError:
    FORCE_OPENPYXL = False

# Configuration de la page
st.set_page_config(
    page_title="G√©n√©rateur ESG - Simple",
    page_icon="üè≠",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Suppression des warnings
warnings.filterwarnings('ignore')

# CSS personnalis√© pour am√©liorer l'apparence avec effets hover
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
        padding: 1rem;
        background: linear-gradient(90deg, #f0f8ff, #e6f3ff);
        border-radius: 10px;
    }
    .step-header {
        font-size: 1.5rem;
        color: #2e8b57;
        margin: 1rem 0;
        padding: 0.5rem;
        background-color: #f0fff0;
        border-left: 4px solid #2e8b57;
        border-radius: 5px;
    }
    .success-box {
        padding: 1rem;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        color: #155724;
        margin: 1rem 0;
    }
    .warning-box {
        padding: 1rem;
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        border-radius: 5px;
        color: #856404;
        margin: 1rem 0;
    }
    .error-box {
        padding: 1rem;
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        border-radius: 5px;
        color: #721c24;
        margin: 1rem 0;
    }
    /* Effets hover pour tous les boutons */
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
        transition: all 0.3s ease;
    }
    .stButton > button {
        transition: all 0.3s ease;
    }
    /* Barre de progression pleine largeur */
    .stProgress > div > div > div > div {
        width: 100% !important;
    }
</style>
""", unsafe_allow_html=True)

def init_session_state():
    """Initialiser les variables de session"""
    if 'temp_dir' not in st.session_state:
        st.session_state.temp_dir = tempfile.mkdtemp()
    if 'uploaded_bdd_file' not in st.session_state:
        st.session_state.uploaded_bdd_file = None
    if 'uploaded_template_file' not in st.session_state:
        st.session_state.uploaded_template_file = None
    if 'selected_year' not in st.session_state:
        st.session_state.selected_year = None
    if 'available_years' not in st.session_state:
        st.session_state.available_years = []
    if 'generated_questionnaires' not in st.session_state:
        st.session_state.generated_questionnaires = False
    if 'current_step' not in st.session_state:
        st.session_state.current_step = 1
    if 'generation_result' not in st.session_state:
        st.session_state.generation_result = None
    if 'generated_zip_data' not in st.session_state:
        st.session_state.generated_zip_data = None
    if 'selected_questionnaire_indices' not in st.session_state:
        st.session_state.selected_questionnaire_indices = []
    if 'questionnaires_preview' not in st.session_state:
        st.session_state.questionnaires_preview = []

def main():
    init_session_state()
    
    # Titre principal
    st.markdown("""
    <div class="main-header">
        üè≠ G√âN√âRATEUR ESG SIMPLIFI√â
        <br><small>G√©n√©ration automatique de questionnaires ESG</small>
    </div>
    """, unsafe_allow_html=True)
    
    # Interface de g√©n√©ration
    page_generation_workflow()

def page_generation_workflow():
    """Interface principale de g√©n√©ration"""
    
    # Navigation par √©tapes
    st.markdown('<div class="step-header">üè≠ G√âN√âRATION DE QUESTIONNAIRES</div>', unsafe_allow_html=True)
    
    # Interface progressive par √©tapes
    if st.session_state.current_step == 1:
        page_upload_files()
    elif st.session_state.current_step == 2:
        page_generation_config()
    elif st.session_state.current_step == 3:
        page_generation_execute()
    
    # Barre de progression avec navigation
    st.markdown("---")
    progress_indicators = st.columns(3)
    with progress_indicators[0]:
        if st.session_state.current_step >= 1:
            if st.session_state.current_step == 1:
                st.warning("üîÑ 1. Upload Fichiers")
            else:
                if st.button("‚úÖ 1. Upload Fichiers", help="Cliquer pour revenir √† cette √©tape"):
                    st.session_state.current_step = 1
                    st.rerun()
        else:
            st.info("‚è∏Ô∏è 1. Upload Fichiers")
    
    with progress_indicators[1]:
        if st.session_state.current_step >= 2:
            if st.session_state.current_step == 2:
                st.warning("üîÑ 2. Configuration")
            else:
                if st.button("‚úÖ 2. Configuration", help="Cliquer pour revenir √† cette √©tape"):
                    st.session_state.current_step = 2
                    st.rerun()
        else:
            st.info("‚è∏Ô∏è 2. Configuration")
    
    with progress_indicators[2]:
        if st.session_state.current_step >= 3:
            st.warning("üîÑ 3. G√©n√©ration")
        else:
            st.info("‚è∏Ô∏è 3. G√©n√©ration")
    
    # Bouton de remise √† z√©ro
    if st.session_state.current_step > 1:
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            if st.button("üîÑ Recommencer", type="secondary", help="Remettre √† z√©ro et recommencer depuis le d√©but"):
                st.session_state.current_step = 1
                st.session_state.uploaded_bdd_file = None
                st.session_state.uploaded_template_file = None
                st.session_state.selected_year = "2025"
                st.session_state.generated_questionnaires = False
                st.session_state.generated_zip_data = None
                st.session_state.selected_questionnaire_indices = []
                st.session_state.questionnaires_preview = []
                st.success("‚úÖ Application remise √† z√©ro")
                st.rerun()

def page_upload_files():
    """Page upload des fichiers"""
    st.markdown("**üìÅ Upload des fichiers requis**")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**1. Fichier BDD ESG**")
        uploaded_bdd = st.file_uploader("S√©lectionnez votre fichier BDD", type=['xlsx', 'xlsm'], key="bdd_upload")
        if uploaded_bdd:
            # Sauvegarder dans le r√©pertoire temporaire
            bdd_path = os.path.join(st.session_state.temp_dir, uploaded_bdd.name)
            with open(bdd_path, "wb") as f:
                f.write(uploaded_bdd.getbuffer())
            st.session_state.uploaded_bdd_file = bdd_path
            st.success("‚úÖ Fichier BDD upload√©")
    
    with col2:
        st.markdown("**2. Template questionnaire**")
        uploaded_template = st.file_uploader("S√©lectionnez votre template", type=['xlsx', 'xlsm'], key="template_upload")
        if uploaded_template:
            # Sauvegarder dans le r√©pertoire temporaire
            template_path = os.path.join(st.session_state.temp_dir, uploaded_template.name)
            with open(template_path, "wb") as f:
                f.write(uploaded_template.getbuffer())
            st.session_state.uploaded_template_file = template_path
            st.success("‚úÖ Template upload√©")
    
    # Analyser le fichier BDD si les deux sont upload√©s
    if st.session_state.uploaded_bdd_file and st.session_state.uploaded_template_file:
        try:
            wb = load_workbook(st.session_state.uploaded_bdd_file, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            # D√©tecter automatiquement les ann√©es (onglets qui sont des nombres de 4 chiffres)
            available_years = []
            for sheet in sheet_names:
                if sheet.isdigit() and len(sheet) == 4 and 2020 <= int(sheet) <= 2030:
                    available_years.append(sheet)
            
            available_years.sort()
            
            if not available_years:
                st.error("‚ùå Aucun onglet d'ann√©e d√©tect√© (format attendu: 2023, 2024, etc.)")
                return
            else:
                st.success(f"‚úÖ Fichiers valid√©s - Ann√©es d√©tect√©es: {', '.join(available_years)}")
                st.session_state.available_years = available_years
                # S√©lectionner la derni√®re ann√©e par d√©faut
                st.session_state.selected_year = available_years[-1]
                
                # Bouton pour passer √† l'√©tape suivante
                if st.button("‚û°Ô∏è Passer √† la configuration", type="primary"):
                    st.session_state.current_step = 2
                    st.rerun()
                    
        except Exception as e:
            st.error(f"‚ùå Erreur lors de l'analyse: {e}")

def page_generation_config():
    """Page configuration pour g√©n√©ration"""
    if not st.session_state.uploaded_bdd_file or not st.session_state.uploaded_template_file:
        st.warning("‚ö†Ô∏è Veuillez d'abord uploader les fichiers")
        return
    
    if not st.session_state.available_years:
        st.error("‚ùå Aucune ann√©e disponible d√©tect√©e")
        return
    
    st.markdown("**‚öôÔ∏è Configuration de la g√©n√©ration**")
    
    # S√©lection de l'ann√©e avec radio buttons bas√© sur les ann√©es d√©tect√©es
    st.markdown("**üìÖ S√©lection de l'ann√©e :**")
    
    # Assurer que selected_year est dans la liste des ann√©es disponibles
    if st.session_state.selected_year not in st.session_state.available_years:
        st.session_state.selected_year = st.session_state.available_years[-1]
    
    selected_year = st.radio(
        "Choisissez l'ann√©e",
        options=st.session_state.available_years,
        index=st.session_state.available_years.index(st.session_state.selected_year),
        horizontal=True,
        key="year_selection"
    )
    
    # Mettre √† jour l'ann√©e s√©lectionn√©e
    if selected_year != st.session_state.selected_year:
        st.session_state.selected_year = selected_year
        st.rerun()
    
    st.write(f"**Ann√©e s√©lectionn√©e :** {st.session_state.selected_year}")
    
    # V√©rification silencieuse de XLWings (pas d'affichage)
    xlwings_gen = XLWingsGenerator()
    xlwings_available, xlwings_msg = xlwings_gen.is_available()
    if not xlwings_available:
        st.error("‚ùå **XLWINGS REQUIS** - V√©rifiez qu'Excel est install√©")
    
    # Pr√©visualisation des donn√©es et s√©lection des questionnaires
    with st.expander("üìä Pr√©visualiser et s√©lectionner les questionnaires √† g√©n√©rer", expanded=True):
        try:
            # Import de la nouvelle fonction
            from generateur_2025_streamlit import get_questionnaires_preview
            
            questionnaires_preview = get_questionnaires_preview(st.session_state.uploaded_bdd_file, st.session_state.selected_year)
            st.session_state.questionnaires_preview = questionnaires_preview
            
            if questionnaires_preview:
                st.write(f"**{len(questionnaires_preview)} questionnaires disponibles pour {st.session_state.selected_year}**")
                
                # Barre de recherche/filtre
                search_term = st.text_input("üîç Rechercher par identifiant, locataire ou adresse :", key="search_filter", placeholder="Tapez votre recherche...")
                
                # Filtrage des questionnaires selon la recherche
                if search_term:
                    search_lower = search_term.lower().strip()
                    filtered_questionnaires = []
                    filtered_indices = []
                    
                    for i, q in enumerate(questionnaires_preview):
                        search_fields = [
                            str(q.get('id', '')).lower(),
                            str(q.get('locataire', '')).lower(), 
                            str(q.get('adresse', '')).lower()
                        ]
                        
                        if any(search_lower in field for field in search_fields):
                            filtered_questionnaires.append(q)
                            filtered_indices.append(i)
                    
                    displayed_questionnaires = filtered_questionnaires
                    displayed_indices = filtered_indices
                    st.info(f"üéØ {len(filtered_questionnaires)} questionnaire(s) trouv√©(s) pour '{search_term}'")
                else:
                    displayed_questionnaires = questionnaires_preview
                    displayed_indices = list(range(len(questionnaires_preview)))
                
                # Boutons de s√©lection rapide (s'appliquent aux r√©sultats filtr√©s)
                col1, col2, col3 = st.columns(3)
                with col1:
                    if st.button("‚úÖ S√©lectionner tout"):
                        # Ajouter les indices affich√©s √† la s√©lection
                        current_selection = set(st.session_state.selected_questionnaire_indices)
                        current_selection.update(displayed_indices)
                        st.session_state.selected_questionnaire_indices = list(current_selection)
                        st.rerun()
                with col2:
                    if st.button("‚ùå D√©s√©lectionner tout"):
                        # Retirer les indices affich√©s de la s√©lection
                        current_selection = set(st.session_state.selected_questionnaire_indices)
                        current_selection -= set(displayed_indices)
                        st.session_state.selected_questionnaire_indices = list(current_selection)
                        st.rerun()
                with col3:
                    if st.button("üîÑ Inverser la s√©lection"):
                        # Inverser uniquement les √©l√©ments affich√©s
                        current_selection = set(st.session_state.selected_questionnaire_indices)
                        displayed_set = set(displayed_indices)
                        # Retirer les s√©lectionn√©s et ajouter les non-s√©lectionn√©s dans les affich√©s
                        to_remove = current_selection & displayed_set
                        to_add = displayed_set - current_selection
                        current_selection -= to_remove
                        current_selection.update(to_add)
                        st.session_state.selected_questionnaire_indices = list(current_selection)
                        st.rerun()
                
                # Affichage des questionnaires avec checkboxes
                st.markdown("**üìã S√©lectionnez les questionnaires √† g√©n√©rer :**")
                
                # Afficher par lots pour √©viter une interface trop lourde
                items_per_page = 20
                total_pages = (len(displayed_questionnaires) + items_per_page - 1) // items_per_page
                
                if total_pages > 1:
                    current_page = st.selectbox("üìÑ Page", range(1, total_pages + 1), index=0) - 1
                else:
                    current_page = 0
                
                start_idx = current_page * items_per_page
                end_idx = min(start_idx + items_per_page, len(displayed_questionnaires))
                
                for display_i in range(start_idx, end_idx):
                    questionnaire = displayed_questionnaires[display_i]
                    original_i = displayed_indices[display_i]  # Index original dans la liste compl√®te
                    
                    # Checkbox pour chaque questionnaire
                    is_selected = original_i in st.session_state.selected_questionnaire_indices
                    
                    col1, col2 = st.columns([1, 4])
                    with col1:
                        selected = st.checkbox("S√©lect.", value=is_selected, key=f"quest_{original_i}_{display_i}", label_visibility="hidden")
                    
                    with col2:
                        # Affichage des informations
                        st.markdown(f"""
                        **#{questionnaire['id']} - {questionnaire['locataire']}**  
                        üìç {questionnaire['adresse']}  
                        üìÖ {questionnaire['date']}  
                        üìÅ `{questionnaire['filename_preview']}`
                        """)
                    
                    # Mettre √† jour la s√©lection
                    if selected and original_i not in st.session_state.selected_questionnaire_indices:
                        st.session_state.selected_questionnaire_indices.append(original_i)
                    elif not selected and original_i in st.session_state.selected_questionnaire_indices:
                        st.session_state.selected_questionnaire_indices.remove(original_i)
                
                # R√©sum√© de la s√©lection
                selected_count = len(st.session_state.selected_questionnaire_indices)
                if selected_count > 0:
                    st.success(f"‚úÖ {selected_count} questionnaire(s) s√©lectionn√©(s) sur {len(questionnaires_preview)}")
                else:
                    st.warning("‚ö†Ô∏è Aucun questionnaire s√©lectionn√©")
                    
            else:
                st.warning(f"Aucune donn√©e trouv√©e pour l'ann√©e {st.session_state.selected_year}")
        except Exception as e:
            st.error(f"Erreur lors de la pr√©visualisation: {e}")
    
    # Bouton pour passer √† l'√©tape suivante
    st.markdown("---")
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        selected_count = len(st.session_state.selected_questionnaire_indices)
        if selected_count > 0:
            if st.button(f"‚úÖ Valider et g√©n√©rer {selected_count} questionnaire(s)", type="primary", use_container_width=True):
                st.success(f"‚úÖ Configuration valid√©e ! {selected_count} questionnaire(s) s√©lectionn√©(s)")
                # Passer automatiquement √† l'√©tape suivante
                st.session_state.current_step = 3
                st.info("‚û°Ô∏è Passage automatique √† l'√©tape G√©n√©ration...")
                time.sleep(1.5)
                st.rerun()
        else:
            st.warning("‚ö†Ô∏è Veuillez s√©lectionner au moins un questionnaire pour continuer")

def page_generation_execute():
    """Page ex√©cution g√©n√©ration"""
    if not st.session_state.uploaded_bdd_file or not st.session_state.uploaded_template_file:
        st.warning("‚ö†Ô∏è Veuillez d'abord uploader les fichiers")
        return
    
    st.markdown("**üöÄ G√©n√©ration des questionnaires**")
    
    # R√©sum√© configuration
    selected_count = len(st.session_state.selected_questionnaire_indices)
    st.markdown(f"""
    **Configuration actuelle :**
    - üìÅ Fichier BDD: `{os.path.basename(st.session_state.uploaded_bdd_file)}`
    - üìã Template: `{os.path.basename(st.session_state.uploaded_template_file)}`
    - üìÖ Ann√©e: `{st.session_state.selected_year}`
    - üìä Questionnaires s√©lectionn√©s: `{selected_count}`
    """)
    
    # Afficher les r√©sultats si la g√©n√©ration a d√©j√† √©t√© faite
    if st.session_state.generated_questionnaires and st.session_state.generation_result:
        display_generation_results()
    else:
        # Bouton de g√©n√©ration avec largeur normale et alignement √† gauche
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            selected_count = len(st.session_state.selected_questionnaire_indices)
            if selected_count > 0:
                if st.button(f"üöÄ G√âN√âRER {selected_count} QUESTIONNAIRE(S)", type="primary"):
                    generate_questionnaires()
            else:
                st.error("‚ùå Aucun questionnaire s√©lectionn√©")

def display_generation_results():
    """Afficher les r√©sultats de g√©n√©ration persistants"""
    result = st.session_state.generation_result
    
    st.success("‚úÖ G√©n√©ration termin√©e avec succ√®s !")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Questionnaires g√©n√©r√©s", result['generated_count'])
    with col2:
        st.metric("√âchecs", result['failed_count'])
    with col3:
        if 'total_available' in result:
            st.metric("Total disponible", result['total_available'])
    
    # Information sur la s√©lection
    if 'total_available' in result and result['total_available'] > result['total_processed']:
        st.info(f"‚ÑπÔ∏è {result['total_processed']} questionnaire(s) s√©lectionn√©(s) sur {result['total_available']} disponibles")
    
    # Bouton de t√©l√©chargement ZIP avec donn√©es directement int√©gr√©es
    if st.session_state.generated_zip_data:
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        selected_count = result.get('total_processed', 'selection')
        zip_filename = f"questionnaires_{st.session_state.selected_year}_{selected_count}items_{timestamp}.zip"
        
        st.download_button(
            label=f"üì• T√©l√©charger le ZIP des questionnaires ({result['generated_count']} fichiers)",
            data=st.session_state.generated_zip_data,
            file_name=zip_filename,
            mime="application/zip",
            type="secondary",
            help=f"T√©l√©charger les {result['generated_count']} questionnaires g√©n√©r√©s pour l'ann√©e {st.session_state.selected_year}"
        )
        
    # Bouton pour recommencer apr√®s g√©n√©ration
    st.markdown("---")
    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        if st.button("üîÑ Nouvelle g√©n√©ration", type="secondary", help="Recommencer une nouvelle g√©n√©ration"):
            st.session_state.current_step = 1
            st.session_state.uploaded_bdd_file = None
            st.session_state.uploaded_template_file = None
            st.session_state.selected_year = "2025"
            st.session_state.generated_questionnaires = False
            st.session_state.generation_result = None
            st.session_state.generated_zip_data = None
            st.session_state.selected_questionnaire_indices = []
            st.session_state.questionnaires_preview = []
            st.success("‚úÖ Pr√™t pour une nouvelle g√©n√©ration")
            st.rerun()

def create_and_show_download_button(output_dir):
    """Cr√©er et afficher le bouton de t√©l√©chargement avec donn√©es int√©gr√©es"""
    try:
        # Nom unique pour le ZIP bas√© sur le timestamp et l'ann√©e
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_filename = f"questionnaires_{st.session_state.selected_year}_{timestamp}.zip"
        zip_path = os.path.join(st.session_state.temp_dir, zip_filename)
        
        # Cr√©er le ZIP avec uniquement les fichiers de cette g√©n√©ration
        if not os.path.exists(zip_path):
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # V√©rifier que le dossier de sortie existe et contient des fichiers
                if os.path.exists(output_dir):
                    files_added = 0
                    for root, dirs, files in os.walk(output_dir):
                        for file in files:
                            if file.endswith(('.xlsx', '.xlsm', '.pdf', '.txt')):  # Filtrer les types de fichiers
                                file_path = os.path.join(root, file)
                                # Nom dans le ZIP relatif au dossier de g√©n√©ration
                                arcname = os.path.relpath(file_path, output_dir)
                                zipf.write(file_path, arcname)
                                files_added += 1
                    
                    if files_added == 0:
                        st.warning("‚ö†Ô∏è Aucun fichier questionnaire trouv√© dans le dossier de g√©n√©ration")
                        return
                    else:
                        st.info(f"üì¶ ZIP cr√©√© avec {files_added} fichiers")
                else:
                    st.error("‚ùå Dossier de g√©n√©ration introuvable")
                    return
        
        # Lire les donn√©es du ZIP
        with open(zip_path, "rb") as f:
            zip_data = f.read()
        
        # Bouton de t√©l√©chargement direct
        st.download_button(
            label=f"üì• T√©l√©charger le ZIP des questionnaires ({st.session_state.selected_year})",
            data=zip_data,
            file_name=zip_filename,
            mime="application/zip",
            type="secondary",
            help=f"T√©l√©charger uniquement les questionnaires g√©n√©r√©s pour l'ann√©e {st.session_state.selected_year}"
        )
            
    except Exception as e:
        st.error(f"‚ùå Erreur lors de la cr√©ation du ZIP: {e}")

def preview_data(bdd_file, year):
    """Pr√©visualiser les donn√©es d'une ann√©e donn√©e"""
    try:
        wb = load_workbook(bdd_file, data_only=True)
        if year not in wb.sheetnames:
            return None
            
        ws = wb[year]
        data = []
        
        for row in range(10, min(50, ws.max_row + 1)):  # Pr√©visualisation limit√©e
            # Colonne O = nom du dossier (colonne 15)
            folder_name = ws.cell(row=row, column=15).value
            if not folder_name or str(folder_name).strip() == '':
                continue
                
            # Autres colonnes importantes
            numero_identification = ws.cell(row=row, column=1).value
            locataire = ws.cell(row=row, column=6).value
            adresse = ws.cell(row=row, column=4).value
            
            data.append({
                'N¬∞ ID': numero_identification,
                'Dossier': str(folder_name).strip(),
                'Locataire': str(locataire).strip() if locataire else '',
                'Adresse': str(adresse).strip() if adresse else ''
            })
        
        wb.close()
        return pd.DataFrame(data)
        
    except Exception as e:
        st.error(f"Erreur lors de la lecture des donn√©es: {e}")
        return None

def generate_questionnaires():
    """G√©n√©rer les questionnaires avec barre de progression et cr√©ation directe du ZIP"""
    
    # Barre de progression pleine largeur
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        # Import du module de g√©n√©ration
        from generateur_2025_streamlit import generate_selected_questionnaires_to_zip
        
        # Param√®tres
        bdd_file = st.session_state.uploaded_bdd_file
        template_file = st.session_state.uploaded_template_file
        year = st.session_state.selected_year
        selected_indices = st.session_state.selected_questionnaire_indices
        
        if not selected_indices:
            st.error("‚ùå Aucun questionnaire s√©lectionn√©")
            return
        
        st.info(f"üìÅ G√©n√©ration de {len(selected_indices)} questionnaire(s) s√©lectionn√©(s) pour l'ann√©e {year}")
        
        # G√©n√©rer directement en ZIP les questionnaires s√©lectionn√©s
        result = generate_selected_questionnaires_to_zip(
            bdd_file, year, template_file, selected_indices,
            progress_callback=lambda current, total, message: update_progress(progress_bar, status_text, current, total, message)
        )
        
        if result['success']:
            st.session_state.generated_questionnaires = True
            st.session_state.generation_result = result
            st.session_state.generated_zip_data = result['zip_data']
            # Recharger la page pour afficher les r√©sultats persistants
            st.rerun()
        else:
            st.error(f"‚ùå Erreur: {result['error']}")
            
    except ImportError as e:
        st.error(f"‚ùå Module de g√©n√©ration non trouv√©: {e}")
        st.error("üîÑ Tentative avec l'ancienne m√©thode...")
        # Fallback vers l'ancienne m√©thode
        generate_questionnaires_fallback()
    except Exception as e:
        st.error(f"‚ùå Erreur inattendue: {str(e)}")

def generate_questionnaires_fallback():
    """M√©thode de g√©n√©ration de fallback"""
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        from generateur_2025_streamlit import generate_questionnaires_for_year
        
        # Param√®tres
        bdd_file = st.session_state.uploaded_bdd_file
        template_file = st.session_state.uploaded_template_file
        year = st.session_state.selected_year
        
        # Dossier de sortie temporaire
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = os.path.join(
            st.session_state.temp_dir,
            f"Questionnaires_{year}_{timestamp}"
        )
        os.makedirs(output_dir, exist_ok=True)
        
        # G√©n√©rer
        result = generate_questionnaires_for_year(
            bdd_file, year, template_file, output_dir,
            progress_callback=lambda current, total, message: update_progress(progress_bar, status_text, current, total, message)
        )
        
        if result['success']:
            # Cr√©er le ZIP en m√©moire
            zip_buffer = create_zip_from_directory(output_dir)
            
            st.session_state.generated_questionnaires = True
            st.session_state.generation_result = result
            st.session_state.generated_zip_data = zip_buffer
            st.rerun()
        else:
            st.error(f"‚ùå Erreur: {result['error']}")
            
    except Exception as e:
        st.error(f"‚ùå Erreur lors du fallback: {str(e)}")

def create_zip_from_directory(directory_path):
    """Cr√©er un ZIP en m√©moire √† partir d'un dossier"""
    import io
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        if os.path.exists(directory_path):
            for root, dirs, files in os.walk(directory_path):
                for file in files:
                    if file.endswith(('.xlsx', '.xlsm')):
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, directory_path)
                        zipf.write(file_path, arcname)
    
    return zip_buffer.getvalue()

def update_progress(progress_bar, status_text, current, total, message):
    """Mettre √† jour la barre de progression"""
    progress = current / total if total > 0 else 0
    progress_bar.progress(progress)
    status_text.text(f"{message} ({current}/{total})")

if __name__ == "__main__":
    main()
