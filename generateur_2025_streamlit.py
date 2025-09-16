#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MODULE GÉNÉRATION QUESTIONNAIRES - Version Streamlit
Adapté de generateur_2025.py pour fonctionner avec Streamlit
Intègre la solution xlwings pour préserver le formatage conditionnel
"""

import os
import shutil
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
import warnings
import zipfile
import io
import tempfile

# Import du générateur xlwings
try:
    from xlwings_generator import XLWingsGenerator
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False
    XLWingsGenerator = None

# FORMULES TEMPLATE INTÉGRÉES - Plus besoin du fichier COPIE.xlsm !
TEMPLATE_FORMULAS = {
    "AA": "='[2]Questionnaire ESG'!$D$105",
    "AB": "='[2]Questionnaire ESG'!$D$115",
    "AC": "='[2]Questionnaire ESG'!$D$118",
    "AD": "='[2]Questionnaire ESG'!$D$129",
    "AE": "='[2]Questionnaire ESG'!$D$132",
    "AF": "='[2]Questionnaire ESG'!$D$142",
    "AG": "='[2]Questionnaire ESG'!$D$145",
    "AH": "='[2]Questionnaire ESG'!$D$156",
    "AI": "='[2]Questionnaire ESG'!$D$159",
    "AJ": "='[2]Questionnaire ESG'!$D$170",
    "AK": "='[2]Questionnaire ESG'!$D$173",
    "AL": "='[2]Questionnaire ESG'!$D$183",
    "AM": "='[2]Questionnaire ESG'!$D$186",
    "AN": "='[2]Questionnaire ESG'!$D$197",
    "AO": "='[2]Questionnaire ESG'!$D$200",
    "AP": "='[2]Questionnaire ESG'!$D$210",
    "AQ": "='[2]Questionnaire ESG'!$D$213",
    "AR": "='[2]Questionnaire ESG'!$D$224",
    "AS": "='[2]Questionnaire ESG'!$D$227",
    "AT": "='[2]Questionnaire ESG'!$D$238",
    "AU": "='[2]Questionnaire ESG'!$D$241",
    "AV": "='[2]Questionnaire ESG'!$D$252",
    "AW": "='[2]Questionnaire ESG'!$D$255",
    "AX": "='[2]Questionnaire ESG'!$D$266",
    "AY": "='[2]Questionnaire ESG'!$D$269",
    "AZ": "='[2]Questionnaire ESG'!$D$280",
    "BA": "='[2]Questionnaire ESG'!$D$283",
    "BB": "='[2]Questionnaire ESG'!$D$296",
    "BC": "='[2]Questionnaire ESG'!$D$299",
    "BD": "='[2]Questionnaire ESG'!$D$315",
    "BE": "='[2]Questionnaire ESG'!$D$318",
    "BF": "='[2]Questionnaire ESG'!$D$328",
    "BG": "='[2]Questionnaire ESG'!$D$331",
    "BH": "='[2]Questionnaire ESG'!$D$347",
    "BI": "='[2]Questionnaire ESG'!$D$350",
    "BJ": "='[2]Questionnaire ESG'!$D$361",
    "BK": "='[2]Questionnaire ESG'!$D$364",
    "BL": "='[2]Questionnaire ESG'!$D$374",
    "BM": "='[2]Questionnaire ESG'!$D$377",
    "BN": "='[2]Questionnaire ESG'!$D$388",
    "BO": "='[2]Questionnaire ESG'!$D$391",
    "BP": "='[2]Questionnaire ESG'!$D$392",
    "BQ": "='[2]Questionnaire ESG'!$D$393",
    "BR": "='[2]Questionnaire ESG'!$D$395",
    "BS": "='[2]Questionnaire ESG'!$D$396",
    "BT": "='[2]Questionnaire ESG'!$D$397",
    "BU": "='[2]Questionnaire ESG'!$D$399",
    "BV": "='[2]Questionnaire ESG'!$D$400",
    "BW": "='[2]Questionnaire ESG'!$D$401",
    "BX": "='[2]Questionnaire ESG'!$D$410",
    "BY": "='[2]Questionnaire ESG'!$D$413",
    "BZ": "='[2]Questionnaire ESG'!$D$427",
    "CA": "='[2]Questionnaire ESG'!$D$430",
    "CB": "='[2]Questionnaire ESG'!$D$431",
    "CC": "='[2]Questionnaire ESG'!$D$432",
    "CD": "='[2]Questionnaire ESG'!$D$434",
    "CE": "='[2]Questionnaire ESG'!$D$435",
    "CF": "='[2]Questionnaire ESG'!$D$436",
    "CG": "='[2]Questionnaire ESG'!$D$438",
    "CH": "='[2]Questionnaire ESG'!$D$439",
    "CI": "='[2]Questionnaire ESG'!$D$440",
    "CJ": "='[2]Questionnaire ESG'!$D$448",
    "CK": "='[2]Questionnaire ESG'!$D$451",
    "CL": "='[2]Questionnaire ESG'!$D$452",
    "CM": "='[2]Questionnaire ESG'!$D$461",
    "CN": "='[2]Questionnaire ESG'!$D$464",
    "CO": "='[2]Questionnaire ESG'!$D$465",
    "CP": "='[2]Questionnaire ESG'!$D$466",
    "CQ": "='[2]Questionnaire ESG'!$D$468",
    "CR": "='[2]Questionnaire ESG'!$D$469",
    "CS": "='[2]Questionnaire ESG'!$D$470",
    "CT": "='[2]Questionnaire ESG'!$D$472",
    "CU": "='[2]Questionnaire ESG'!$D$473",
    "CV": "='[2]Questionnaire ESG'!$D$474",
    "CW": "='[2]Questionnaire ESG'!$D$476",
    "CX": "='[2]Questionnaire ESG'!$D$477",
    "CY": "='[2]Questionnaire ESG'!$D$478",
    "CZ": "='[2]Questionnaire ESG'!$D$487",
    "DA": "='[2]Questionnaire ESG'!$D$490",
    "DB": "='[2]Questionnaire ESG'!$D$501",
    "DC": "='[2]Questionnaire ESG'!$D$504",
    "DD": "='[2]Questionnaire ESG'!$D$514",
    "DE": "='[2]Questionnaire ESG'!$D$517",
    "DF": "='[2]Questionnaire ESG'!$D$528",
    "DG": "='[2]Questionnaire ESG'!$D$531",
    "DH": "='[2]Questionnaire ESG'!$D$541",
    "DI": "='[2]Questionnaire ESG'!$D$544",
    "DJ": "='[2]Questionnaire ESG'!$D$554",
    "DK": "='[2]Questionnaire ESG'!$D$557",
    "DL": "='[2]Questionnaire ESG'!$D$567",
    "DM": "='[2]Questionnaire ESG'!$D$570",
    "DN": "='[2]Questionnaire ESG'!$D$580",
    "DO": "='[2]Questionnaire ESG'!$D$583",
    "J": "='[2]Questionnaire ESG'!$B$14",
    "K": "='[2]Questionnaire ESG'!$B$17",
    "L": "='[2]Questionnaire ESG'!$B$20",
    "P": "='[2]Questionnaire ESG'!$D$33",
    "Q": "='[2]Questionnaire ESG'!$D$36",
    "R": "='[2]Questionnaire ESG'!$D$46",
    "S": "='[2]Questionnaire ESG'!$D$49",
    "T": "='[2]Questionnaire ESG'!$D$60",
    "U": "='[2]Questionnaire ESG'!$D$63",
    "V": "='[2]Questionnaire ESG'!$D$74",
    "W": "='[2]Questionnaire ESG'!$D$77",
    "X": "='[2]Questionnaire ESG'!$D$88",
    "Y": "='[2]Questionnaire ESG'!$D$91",
    "Z": "='[2]Questionnaire ESG'!$D$102",
}

def extract_cell_references(formula):
    """Extrait les références de cellules d'une formule Excel"""
    if not formula or not formula.startswith('='):
        return None
    
    external_ref_pattern = r"'\[([^\]]+)\]([^']+)'\!(\$?[A-Z]+\$?[0-9]+)"
    match = re.search(external_ref_pattern, formula)
    if match:
        sheet_name = match.group(2)
        cell_ref = match.group(3).replace('$', '')
        return (sheet_name, cell_ref)
    return None

def read_year_data(bdd_file, year):
    """Lire les données de l'onglet spécifié"""
    wb = load_workbook(bdd_file, data_only=True)
    if year not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Onglet {year} non trouvé!")
    
    ws = wb[year]
    data = []
    
    # Lire à partir de la ligne 10
    max_row = 200
    for row in range(10, max_row):
        # Colonne O = nom du dossier (colonne 15)
        folder_name = ws.cell(row=row, column=15).value
        if not folder_name or str(folder_name).strip() == '':
            continue
        
        # Colonne A = Numéro d'identification (colonne 1)
        numero_identification = ws.cell(row=row, column=1).value
        
        # Colonne F = Locataire (colonne 6)
        locataire = ws.cell(row=row, column=6).value
        
        # Colonne D = Adresse (colonne 4)
        adresse_complete = ws.cell(row=row, column=4).value
        
        # Colonne M = Réponse certifiée (colonne 13)
        reponse_certifiee = ws.cell(row=row, column=13).value
        
        # Colonne L = Date (colonne 12) - NOUVELLE LECTURE
        date_questionnaire = ws.cell(row=row, column=12).value
        
        # Extraire ville et adresse du nom du dossier
        folder_parts = str(folder_name).strip().split(' - ')
        ville = folder_parts[0] if len(folder_parts) > 0 else 'VILLE'
        adresse = folder_parts[1] if len(folder_parts) > 1 else 'ADRESSE'
        locataire_final = str(locataire).strip() if locataire else 'LOCATAIRE'
        
        row_data = {
            'row_number': row,
            'numero_identification': str(numero_identification).strip() if numero_identification else str(row-9),
            'folder_name': str(folder_name).strip(),
            'ville': ville,
            'adresse': adresse,
            'locataire': locataire_final,
            'adresse_complete': str(adresse_complete).strip() if adresse_complete else adresse,
            'reponse_certifiee': str(reponse_certifiee).strip() if reponse_certifiee else '',
            'date_questionnaire': date_questionnaire  # NOUVELLE DONNÉE
        }
        
        data.append(row_data)
    
    wb.close()
    return data

def generate_single_questionnaire(template_file, output_folder, data_row, source_file, year):
    """Générer un questionnaire pré-rempli pour un locataire"""
    
    # Créer le dossier du locataire
    folder_structure = f"{data_row['ville']} - {data_row['adresse']} - {data_row['locataire']}"
    final_output_folder = os.path.join(output_folder, folder_structure)
    
    if not os.path.exists(final_output_folder):
        os.makedirs(final_output_folder)
    
    # Nom du fichier de sortie avec le nouveau format
    
    # 1. Récupérer la date depuis la colonne L de la BDD
    date_questionnaire = data_row.get('date_questionnaire')
    if date_questionnaire and hasattr(date_questionnaire, 'strftime'):
        date_formatted = date_questionnaire.strftime("%Y %m %d")
    elif date_questionnaire:
        # Essayer de parser la date si c'est une string
        try:
            if isinstance(date_questionnaire, str):
                # Essayer différents formats
                for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"]:
                    try:
                        parsed_date = datetime.datetime.strptime(date_questionnaire, fmt)
                        date_formatted = parsed_date.strftime("%Y %m %d")
                        break
                    except:
                        continue
                else:
                    # Si aucun format ne marche, utiliser la date actuelle
                    date_formatted = datetime.datetime.now().strftime("%Y %m %d")
            else:
                date_formatted = datetime.datetime.now().strftime("%Y %m %d")
        except:
            date_formatted = datetime.datetime.now().strftime("%Y %m %d")
    else:
        # Si pas de date, utiliser la date actuelle
        date_formatted = datetime.datetime.now().strftime("%Y %m %d")
    
    # 2. Nouveau format : "AAAA MM JJ – Questionnaire ESG – Rue, n° - locataire - #id"
    rue_numero = data_row['adresse_complete']
    locataire = data_row['locataire']
    numero_id = data_row['numero_identification']
    
    output_filename = f"{date_formatted} – Questionnaire ESG – {rue_numero} - {locataire} - #{numero_id}.xlsx"
    output_path = os.path.join(final_output_folder, output_filename)
    
    try:
        # Copier le template
        shutil.copy2(template_file, output_path)
        
        # Ouvrir avec openpyxl pour pré-remplir
        wb = load_workbook(output_path)
        
        # Pré-remplir les données spéciales
        if 'Questionnaire ESG' in wb.sheetnames:
            ws_questionnaire = wb['Questionnaire ESG']
            try:
                # Locataire en B11
                ws_questionnaire['B11'] = data_row['locataire']
                # Adresse en B8
                ws_questionnaire['B8'] = data_row['adresse_complete']
                # Réponse certifiée en A22
                if data_row['reponse_certifiee']:
                    ws_questionnaire['A22'] = data_row['reponse_certifiee']
            except Exception:
                pass
        
        # Utiliser le système de mapping des formules
        wb_source = load_workbook(source_file, data_only=True)
        ws_source = wb_source[year]
        
        updates_count = 0
        for col_letter, formula in TEMPLATE_FORMULAS.items():
            cell_info = extract_cell_references(formula)
            
            if cell_info:
                sheet_name, cell_ref = cell_info
                try:
                    # Lire la valeur depuis la source
                    col_num = column_index_from_string(col_letter)
                    source_value = ws_source.cell(row=data_row['row_number'], column=col_num).value
                    
                    if source_value is not None and str(source_value).strip() != '':
                        # Écrire dans le questionnaire
                        if sheet_name in wb.sheetnames:
                            ws_target = wb[sheet_name]
                            ws_target[cell_ref] = source_value
                            updates_count += 1
                        
                except Exception:
                    continue
        
        wb_source.close()
        
        # Sauvegarder
        wb.save(output_path)
        wb.close()
        
        return True, updates_count
        
    except Exception as e:
        return False, str(e)

def generate_questionnaires_for_year(bdd_file, year, template_file, output_dir, progress_callback=None):
    """Générer tous les questionnaires pour une année donnée"""
    
    try:
        warnings.filterwarnings('ignore')
        
        # Lire les données de l'année
        data_list = read_year_data(bdd_file, year)
        
        if not data_list:
            return {'success': False, 'error': f'Aucune donnée trouvée pour l\'année {year}'}
        
        generated_count = 0
        failed_count = 0
        
        for i, data in enumerate(data_list):
            if progress_callback:
                progress_callback(i, len(data_list), f"Génération {data['ville']} - {data['adresse']} - {data['locataire']}")
            
            success, result = generate_single_questionnaire(
                template_file, output_dir, data, bdd_file, year
            )
            
            if success:
                generated_count += 1
            else:
                failed_count += 1
        
        # Progression finale
        if progress_callback:
            progress_callback(len(data_list), len(data_list), "Génération terminée")
        
        return {
            'success': True,
            'generated_count': generated_count,
            'failed_count': failed_count,
            'total_processed': len(data_list)
        }
        
    except Exception as e:
        return {'success': False, 'error': str(e)}


def generate_questionnaires_for_year_to_zip(bdd_file, year, template_file, progress_callback=None):
    """
    Générer les questionnaires directement en mémoire et retourner un ZIP
    
    Args:
        bdd_file (str): Chemin du fichier BDD
        year (int): Année de génération
        template_file (str): Chemin du fichier template
        progress_callback (function): Fonction de rappel pour la progression
    
    Returns:
        dict: Résultat avec 'success', 'zip_data', et statistiques
    """
    import zipfile
    import io
    import tempfile
    
    try:
        # Créer un buffer ZIP en mémoire
        zip_buffer = io.BytesIO()
        
        # Lire les données avec la fonction existante
        data_list = read_year_data(bdd_file, year)
        
        if not data_list:
            return {'success': False, 'error': f'Aucune donnée trouvée pour l\'année {year}'}
        
        generated_count = 0
        failed_count = 0
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for i, data in enumerate(data_list):
                if progress_callback:
                    progress_callback(i, len(data_list), f"Génération {data['ville']} - {data['adresse']} - {data['locataire']}")
                
                try:
                    # Créer un fichier temporaire pour ce questionnaire
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_path = temp_file.name
                    
                    # Générer le questionnaire dans le fichier temporaire
                    success, result = generate_single_questionnaire_to_file(
                        template_file, temp_path, data, bdd_file, year
                    )
                    
                    if success:
                        # Créer la structure de dossiers comme dans l'ancienne version
                        folder_structure = f"{data['ville']} - {data['adresse']} - {data['locataire']}"
                        filename = create_filename_with_folder(data, year, folder_structure)
                        zipf.write(temp_path, filename)
                        generated_count += 1
                    else:
                        failed_count += 1
                    
                    # Nettoyer le fichier temporaire
                    try:
                        os.unlink(temp_path)
                    except:
                        pass
                        
                except Exception as e:
                    failed_count += 1
                    print(f"Erreur lors de la génération pour {data.get('locataire', 'Unknown')}: {e}")
        
        # Progression finale
        if progress_callback:
            progress_callback(len(data_list), len(data_list), "Génération terminée")
        
        # Retourner les données ZIP
        zip_data = zip_buffer.getvalue()
        
        return {
            'success': True,
            'zip_data': zip_data,
            'generated_count': generated_count,
            'failed_count': failed_count,
            'total_processed': len(data_list)
        }
        
    except Exception as e:
        return {'success': False, 'error': str(e)}


def generate_single_questionnaire_to_file(template_file, output_path, data, bdd_file, year, enable_vba_formatting=True, enable_xlwings=True):
    """
    Générer un questionnaire vers un fichier spécifique
    Version avec support xlwings pour préservation parfaite du formatage
    
    Args:
        template_file: Chemin vers le template Excel
        output_path: Chemin de sortie souhaité  
        data: Données du questionnaire
        bdd_file: Fichier de base de données
        year: Année
        enable_vba_formatting: Activer le formatage VBA/openpyxl
        enable_xlwings: Activer xlwings (priorité si disponible)
    """
    
    # UTILISATION EXCLUSIVE DE XLWINGS - PLUS DE FALLBACK OPENPYXL
    if not enable_xlwings:
        error_msg = "❌ XLWINGS REQUIS: XLWings est désactivé mais requis pour cette application"
        print(error_msg)
        return False, error_msg
    
    if not XLWINGS_AVAILABLE:
        error_msg = "❌ XLWINGS REQUIS: Module XLWings non disponible"
        print(error_msg)
        return False, error_msg
        
    xlwings_gen = XLWingsGenerator()
    available, status = xlwings_gen.is_available()
    
    if not available:
        error_msg = f"❌ XLWINGS REQUIS: {status}"
        print(error_msg)
        return False, error_msg
    
    # Génération avec XLWings uniquement
    try:
        # Préparer les données pour xlwings
        xlwings_data = prepare_data_for_xlwings(data, bdd_file, year)
        
        # Générer avec xlwings
        result = xlwings_gen.generate_single_questionnaire(
            template_file, 
            xlwings_data, 
            output_path
        )
        
        if result["success"]:
            print(f"✅ Questionnaire généré avec xlwings: {output_path}")
            return True, result.get("message", "Questionnaire généré avec xlwings")
        else:
            error_msg = f"❌ XLWINGS ÉCHEC: {result['error']}"
            print(error_msg)
            return False, error_msg
            
    except Exception as e:
        error_msg = f"❌ ERREUR XLWINGS: {str(e)}"
        print(error_msg)
        return False, error_msg


def prepare_data_for_xlwings(data, bdd_file, year):
    """
    Prépare les données dans le format attendu par xlwings
    
    Args:
        data: Données originales
        bdd_file: Fichier BDD
        year: Année
        
    Returns:
        Dict avec les données formatées pour xlwings
    """
    try:
        # Charger les données de la BDD pour cette ligne
        wb_source = load_workbook(bdd_file, data_only=True)
        ws_source = wb_source[year]
        
        # Extraire toutes les données de la ligne
        row_data = {}
        
        # Données de base
        row_data.update({
            'nom_locataire': data.get('locataire', ''),
            'adresse': data.get('adresse_complete', ''),
            'ville': data.get('ville', ''),
            'code_postal': data.get('code_postal', ''),
            'contact': data.get('contact', ''),
            'telephone': data.get('telephone', ''),
            'email': data.get('email', ''),
            'date': datetime.datetime.now().strftime('%Y %m %d'),
            'reference': data.get('reference', ''),
            'reponse_certifiee': data.get('reponse_certifiee', '')
        })
        
        # Extraire les scores et données de la BDD
        if 'row_number' in data:
            row_num = data['row_number']
            
            # Parcourir toutes les colonnes pour extraire les données
            for col_idx in range(1, ws_source.max_column + 1):
                try:
                    cell_value = ws_source.cell(row=row_num, column=col_idx).value
                    col_letter = get_column_letter(col_idx)
                    row_data[f'col_{col_letter}'] = cell_value
                except:
                    continue
        
        wb_source.close()
        return row_data
        
    except Exception as e:
        print(f"Erreur préparation données xlwings: {e}")
        # Retourner au minimum les données de base
        return {
            'nom_locataire': data.get('locataire', ''),
            'adresse': data.get('adresse_complete', ''),
            'date': datetime.datetime.now().strftime('%Y %m %d')
        }


def generate_single_questionnaire_to_file_openpyxl(template_file, output_path, data, bdd_file, year, enable_vba_formatting=True):
    """
    Générer un questionnaire avec openpyxl (méthode classique)
    Version renommée de la fonction originale
    """
    try:
        # Copier le template vers le fichier de sortie
        shutil.copy2(template_file, output_path)
        
        # Charger et modifier le workbook
        wb = load_workbook(output_path)
        
        # Pré-remplir les données spéciales
        if 'Questionnaire ESG' in wb.sheetnames:
            ws_questionnaire = wb['Questionnaire ESG']
            try:
                # Locataire en B11
                ws_questionnaire['B11'] = data['locataire']
                # Adresse en B8
                ws_questionnaire['B8'] = data['adresse_complete']
                # Réponse certifiée en A22
                if data['reponse_certifiee']:
                    ws_questionnaire['A22'] = data['reponse_certifiee']
            except Exception:
                pass
        
        # Utiliser le système de mapping des formules
        wb_source = load_workbook(bdd_file, data_only=True)
        ws_source = wb_source[year]
        
        updates_count = 0
        for col_letter, formula in TEMPLATE_FORMULAS.items():
            cell_info = extract_cell_references(formula)
            
            if cell_info:
                sheet_name, cell_ref = cell_info
                try:
                    # Lire la valeur depuis la source
                    col_num = column_index_from_string(col_letter)
                    source_value = ws_source.cell(row=data['row_number'], column=col_num).value
                    
                    if source_value is not None and str(source_value).strip() != '':
                        # Écrire dans le questionnaire
                        if sheet_name in wb.sheetnames:
                            ws_target = wb[sheet_name]
                            ws_target[cell_ref] = source_value
                            updates_count += 1
                        
                except Exception:
                    continue
        
        wb_source.close()
        wb.save(output_path)
        wb.close()
        
        return True, f"Questionnaire généré avec {updates_count} valeurs (openpyxl)"
        
    except Exception as e:
        return False, f"Erreur lors de la génération: {str(e)}"
    try:
        # Copier le template vers le fichier de sortie
        shutil.copy2(template_file, output_path)
        
        # Charger et modifier le workbook
        wb = load_workbook(output_path)
        
        # Pré-remplir les données spéciales
        if 'Questionnaire ESG' in wb.sheetnames:
            ws_questionnaire = wb['Questionnaire ESG']
            try:
                # Locataire en B11
                ws_questionnaire['B11'] = data['locataire']
                # Adresse en B8
                ws_questionnaire['B8'] = data['adresse_complete']
                # Réponse certifiée en A22
                if data['reponse_certifiee']:
                    ws_questionnaire['A22'] = data['reponse_certifiee']
            except Exception:
                pass
        
        # Utiliser le système de mapping des formules
        wb_source = load_workbook(bdd_file, data_only=True)
        ws_source = wb_source[year]
        
        updates_count = 0
        for col_letter, formula in TEMPLATE_FORMULAS.items():
            cell_info = extract_cell_references(formula)
            
            if cell_info:
                sheet_name, cell_ref = cell_info
                try:
                    # Lire la valeur depuis la source
                    col_num = column_index_from_string(col_letter)
                    source_value = ws_source.cell(row=data['row_number'], column=col_num).value
                    
                    if source_value is not None and str(source_value).strip() != '':
                        # Écrire dans le questionnaire
                        if sheet_name in wb.sheetnames:
                            ws_target = wb[sheet_name]
                            ws_target[cell_ref] = source_value
                            updates_count += 1
                        
                except Exception:
                    continue
        
        wb_source.close()
        
        # Sauvegarder
        wb.save(output_path)
        wb.close()
        
        return True, f"Questionnaire généré avec {updates_count} valeurs"
        
    except Exception as e:
        return False, f"Erreur lors de la génération: {str(e)}"


def create_filename_with_folder(data, year, folder_structure):
    """Créer un nom de fichier avec la structure de dossiers pour le ZIP"""
    try:
        # Nettoyer les caractères spéciaux pour le dossier
        def clean_name(text):
            if not text:
                return "Unknown"
            # Remplacer les caractères interdits dans les noms de fichiers
            text = str(text).strip()
            text = re.sub(r'[<>:"/\\|?*]', '_', text)
            text = re.sub(r'\s+', ' ', text)
            return text[:50]  # Limiter la longueur
        
        # Nettoyer le nom du dossier
        clean_folder = clean_name(folder_structure)
        
        # Nouveau format : "AAAA MM JJ – Questionnaire ESG – Rue, n° - locataire - #id"
        
        # 1. Récupérer la date depuis la colonne L de la BDD
        date_questionnaire = data.get('date_questionnaire')
        if date_questionnaire and hasattr(date_questionnaire, 'strftime'):
            date_formatted = date_questionnaire.strftime("%Y %m %d")
        elif date_questionnaire:
            # Essayer de parser la date si c'est une string
            try:
                if isinstance(date_questionnaire, str):
                    # Essayer différents formats
                    for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"]:
                        try:
                            parsed_date = datetime.datetime.strptime(date_questionnaire, fmt)
                            date_formatted = parsed_date.strftime("%Y %m %d")
                            break
                        except:
                            continue
                    else:
                        # Si aucun format ne marche, utiliser la date actuelle
                        date_formatted = datetime.datetime.now().strftime("%Y %m %d")
                else:
                    date_formatted = datetime.datetime.now().strftime("%Y %m %d")
            except:
                date_formatted = datetime.datetime.now().strftime("%Y %m %d")
        else:
            # Si pas de date, utiliser la date actuelle
            date_formatted = datetime.datetime.now().strftime("%Y %m %d")
        
        # 2. Préparer les autres éléments
        numero = data.get('numero_identification', '00')
        
        # Extraire rue et numéro de l'adresse complète
        adresse_complete = data.get('adresse_complete', data.get('adresse', 'Unknown'))
        rue_numero = clean_name(adresse_complete)
        
        locataire = clean_name(data.get('locataire', 'Unknown'))
        
        # 3. Construire le nom selon le format demandé
        filename = f"{date_formatted} – Questionnaire ESG – {rue_numero} - {locataire} - #{numero}.xlsx"
        
        # S'assurer que le nom n'est pas trop long
        if len(filename) > 200:
            filename = f"{date_formatted} – Questionnaire ESG – {locataire} - #{numero}.xlsx"
        
        # Retourner le chemin complet : dossier/fichier
        return f"{clean_folder}/{filename}"
        
    except Exception as e:
        # Fallback en cas d'erreur
        folder_fallback = f"Questionnaire_{year}"
        date_fallback = datetime.datetime.now().strftime("%Y %m %d")
        numero_fallback = data.get('numero_identification', 'unknown')
        filename_fallback = f"{date_fallback} – Questionnaire ESG – #{numero_fallback}.xlsx"
        return f"{folder_fallback}/{filename_fallback}"


def create_filename(data, year):
    """Créer un nom de fichier pour le questionnaire (version simple sans dossier)"""
    try:
        # Nettoyer les caractères spéciaux
        def clean_name(text):
            if not text:
                return "Unknown"
            # Remplacer les caractères interdits dans les noms de fichiers
            text = str(text).strip()
            text = re.sub(r'[<>:"/\\|?*]', '_', text)
            text = re.sub(r'\s+', ' ', text)
            return text[:50]  # Limiter la longueur
        
        ville = clean_name(data.get('ville', 'Unknown'))
        adresse = clean_name(data.get('adresse', 'Unknown'))
        locataire = clean_name(data.get('locataire', 'Unknown'))
        numero = data.get('numero_identification', data.get('numero', '00'))
        
        date_generation = datetime.datetime.now().strftime("%Y %m %d")
        filename = f"#{numero} - {date_generation} - Questionnaire ESG - {ville} - {adresse} - {locataire}.xlsx"
        
        # S'assurer que le nom n'est pas trop long
        if len(filename) > 200:
            filename = f"#{numero} - {date_generation} - {locataire}.xlsx"
        
        return filename
        
    except Exception as e:
        return f"Questionnaire_{year}_{data.get('numero_identification', 'unknown')}.xlsx"


def get_questionnaires_preview(bdd_file, year):
    """
    Obtenir une prévisualisation des questionnaires disponibles pour sélection
    
    Args:
        bdd_file (str): Chemin du fichier BDD
        year (str): Année à traiter
        
    Returns:
        list: Liste des questionnaires avec infos pour affichage
    """
    try:
        data_list = read_year_data(bdd_file, year)
        preview_list = []
        
        for i, data in enumerate(data_list):
            # Formatter la date pour l'affichage
            date_questionnaire = data.get('date_questionnaire')
            if date_questionnaire and hasattr(date_questionnaire, 'strftime'):
                date_str = date_questionnaire.strftime("%Y-%m-%d")
            elif date_questionnaire:
                date_str = str(date_questionnaire)
            else:
                date_str = "Date non définie"
            
            # Créer un aperçu du nom de fichier qui sera généré
            if date_questionnaire and hasattr(date_questionnaire, 'strftime'):
                date_formatted = date_questionnaire.strftime("%Y %m %d")
            else:
                date_formatted = datetime.datetime.now().strftime("%Y %m %d")
            
            preview_filename = f"{date_formatted} – Questionnaire ESG – {data['adresse_complete']} - {data['locataire']} - #{data['numero_identification']}.xlsx"
            
            preview_item = {
                'index': i,
                'id': data['numero_identification'],
                'locataire': data['locataire'],
                'adresse': data['adresse_complete'],
                'date': date_str,
                'filename_preview': preview_filename,
                'data': data  # Données complètes pour la génération
            }
            preview_list.append(preview_item)
        
        return preview_list
        
    except Exception as e:
        return []


def generate_selected_questionnaires_to_zip(bdd_file, year, template_file, selected_indices, progress_callback=None):
    """
    Générer uniquement les questionnaires sélectionnés vers un ZIP
    
    Args:
        bdd_file (str): Chemin du fichier BDD
        year (str): Année à traiter
        template_file (str): Chemin du template
        selected_indices (list): Liste des indices des questionnaires à générer
        progress_callback (function): Callback pour la progression
        
    Returns:
        dict: Résultat avec zip_data et statistiques
    """
    import zipfile
    import io
    import tempfile
    
    try:
        # Obtenir toutes les données
        all_data = read_year_data(bdd_file, year)
        
        if not all_data:
            return {'success': False, 'error': f'Aucune donnée trouvée pour l\'année {year}'}
        
        # Filtrer selon les indices sélectionnés
        selected_data = [all_data[i] for i in selected_indices if i < len(all_data)]
        
        if not selected_data:
            return {'success': False, 'error': 'Aucun questionnaire sélectionné valide'}
        
        # Créer un buffer ZIP en mémoire
        zip_buffer = io.BytesIO()
        
        generated_count = 0
        failed_count = 0
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for i, data in enumerate(selected_data):
                if progress_callback:
                    progress_callback(i, len(selected_data), f"Génération {data['locataire']}")
                
                try:
                    # Créer un fichier temporaire pour ce questionnaire
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_path = temp_file.name
                    
                    # Générer le questionnaire dans le fichier temporaire
                    success, result = generate_single_questionnaire_to_file(
                        template_file, temp_path, data, bdd_file, year
                    )
                    
                    if success:
                        # Créer la structure de dossiers comme dans l'ancienne version
                        folder_structure = f"{data['ville']} - {data['adresse']} - {data['locataire']}"
                        filename = create_filename_with_folder(data, year, folder_structure)
                        zipf.write(temp_path, filename)
                        generated_count += 1
                    else:
                        failed_count += 1
                    
                    # Nettoyer le fichier temporaire
                    try:
                        os.unlink(temp_path)
                    except:
                        pass
                        
                except Exception as e:
                    failed_count += 1
                    print(f"Erreur lors de la génération pour {data.get('locataire', 'Unknown')}: {e}")
        
        # Progression finale
        if progress_callback:
            progress_callback(len(selected_data), len(selected_data), "Génération terminée")
        
        # Retourner les données ZIP
        zip_data = zip_buffer.getvalue()
        
        return {
            'success': True,
            'zip_data': zip_data,
            'generated_count': generated_count,
            'failed_count': failed_count,
            'total_processed': len(selected_data),
            'total_available': len(all_data)
        }
        
    except Exception as e:
        return {'success': False, 'error': str(e)}
